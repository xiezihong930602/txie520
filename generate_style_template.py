"""
生成款式库Excel模板
基于重量表 + 产品目录表整合
"""
from openpyxl import Workbook, load_workbook
import os

os.chdir('C:/Users/Administrator/.doubao/chats/2026-07-14/new-chat/temu_auto_publish_v2')

# 读取重量表
wb_weight = load_workbook('服装重量.xlsx', data_only=True)

def parse_weight_sheet(sheet_name, category):
    """解析重量表，返回款式列表"""
    ws = wb_weight[sheet_name]
    styles = []
    
    # 读取表头（尺码行）
    header = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 1:  # 第2行是表头
            header = row
            break
    
    # 尺码列从第3列开始
    size_cols = []
    for idx, val in enumerate(header):
        if idx >= 2 and val and str(val).strip() and str(val).strip() != 'None':
            size_cols.append((idx, str(val).strip()))
    
    # 读取款式数据
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        name = row[0]
        if not name or str(name).strip() in ['', '商品', '产品款名']:
            continue
        
        style_name = str(name).strip()
        color = row[1] if len(row) > 1 else None
        
        # 提取各尺码重量
        size_weights = {}
        sizes = []
        for col_idx, size_name in size_cols:
            weight = row[col_idx] if col_idx < len(row) else None
            if weight and str(weight).strip() and str(weight).strip() != '-':
                try:
                    w = float(weight)
                    size_weights[size_name] = w
                    sizes.append(size_name)
                except:
                    pass
        
        if not sizes:
            continue
        
        # 计算参考重量（取中间尺码）
        mid_idx = len(sizes) // 2
        ref_weight = size_weights[sizes[mid_idx]]
        
        styles.append({
            'name': style_name,
            'category': category,
            'sizes': sizes,
            'size_weights': size_weights,
            'ref_weight': ref_weight,
            'color': color
        })
    
    return styles

# 读取儿童款和成人款
styles_kids = parse_weight_sheet('儿童', '儿童')
styles_adult = parse_weight_sheet('成人', '成人')
all_styles = styles_kids + styles_adult

print(f'解析到 {len(all_styles)} 款（儿童{len(styles_kids)} + 成人{len(styles_adult)}）')

# 读取产品目录表，尝试匹配成本价
wb_cat = load_workbook('产品目录表.xlsx', data_only=True, read_only=True)

def parse_cat_sheet(sheet_name):
    ws = wb_cat[sheet_name]
    items = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        name = row[1]
        if not name or str(name).strip() in ['', '产品款名']:
            continue
        clean_name = str(name).strip().replace('\n', '')
        cost = row[5]
        fabric = row[4]
        items.append({
            'name': clean_name,
            'cost': cost,
            'fabric': fabric
        })
    return items

cat_kids = parse_cat_sheet('儿童款')
cat_adult = parse_cat_sheet('成人')
all_cat = cat_kids + cat_adult

# 简单匹配成本（基于关键词，仅供参考）
def match_cost(style_name, cat_list):
    best_match = None
    best_score = 0
    s_clean = style_name.replace(' ', '')
    
    for cat in cat_list:
        c_clean = cat['name'].replace(' ', '')
        common = len(set(s_clean) & set(c_clean))
        score = common / max(len(s_clean), len(c_clean))
        if score > best_score:
            best_score = score
            best_match = cat
    
    if best_score > 0.45:
        return best_match['cost'], best_match['fabric']
    return '', ''

# 生成Excel
wb_out = Workbook()

# Sheet1: 款式库主表
ws_main = wb_out.active
ws_main.title = '款式库主表'

# 表头
headers = [
    '款式编号', '款式名称', '分类', '品类模板',
    '成本价(元)', '面料', '季节', '版型', '织造方式', '场景', '场合',
    '默认SKU分类', '单件净重参考(g)', '单件毛重(g)',
    '尺码列表', '备注'
]
ws_main.append(headers)

# 加粗表头
from openpyxl.styles import Font
for col in range(1, len(headers) + 1):
    ws_main.cell(row=1, column=col).font = Font(bold=True)

# 填充数据
for idx, style in enumerate(all_styles, 1):
    style_id = f"ST-{idx:03d}"
    
    # 匹配成本
    cat_list = cat_kids if style['category'] == '儿童' else cat_adult
    cost, fabric = match_cost(style['name'], cat_list)
    
    sizes_str = ','.join(style['sizes'])
    
    row_data = [
        style_id,
        style['name'],
        style['category'],
        '',  # 品类模板
        cost,
        fabric,
        '',  # 季节
        '',  # 版型
        '',  # 织造方式
        '',  # 场景
        '',  # 场合
        '',  # SKU分类
        style['ref_weight'],
        '',  # 毛重
        sizes_str,
        ''   # 备注
    ]
    ws_main.append(row_data)

# 调整列宽
col_widths = [12, 28, 8, 16, 12, 14, 8, 8, 12, 8, 8, 12, 14, 12, 30, 20]
for i, w in enumerate(col_widths, 1):
    ws_main.column_dimensions[chr(64 + i)].width = w

# Sheet2: 尺码重量明细表
ws_detail = wb_out.create_sheet('尺码重量明细')
ws_detail.append(['款式编号', '款式名称', '分类', '尺码', '净重(g)'])
for col in range(1, 6):
    ws_detail.cell(row=1, column=col).font = Font(bold=True)

for idx, style in enumerate(all_styles, 1):
    style_id = f"ST-{idx:03d}"
    for size, weight in style['size_weights'].items():
        ws_detail.append([
            style_id,
            style['name'],
            style['category'],
            size,
            weight
        ])

# 调整列宽
detail_widths = [12, 28, 8, 10, 12]
for i, w in enumerate(detail_widths, 1):
    ws_detail.column_dimensions[chr(64 + i)].width = w

# 保存
output_file = '款式库模板.xlsx'
wb_out.save(output_file)
print(f'已生成 {output_file}')
print(f'主表 {len(all_styles)} 行，明细表 {sum(len(s["sizes"]) for s in all_styles)} 行')
