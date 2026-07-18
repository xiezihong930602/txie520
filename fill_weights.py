"""
任务1：把服装重量表的净重数据填入款式库模板的尺码明细
任务2：调整尺码明细的字段顺序
"""
from openpyxl import load_workbook
from openpyxl.styles import Font
import os

os.chdir('C:/Users/Administrator/.doubao/chats/2026-07-14/new-chat/temu_auto_publish_v2')

# ===== 1. 读取重量表，建立映射 =====
wb_weight = load_workbook('服装重量_最新版.xlsx', data_only=True)

def build_weight_map(sheet_name):
    """读取重量表sheet，返回 {款式名: {尺码: 重量}}"""
    ws = wb_weight[sheet_name]
    weight_map = {}
    
    # 读取表头（第2行）
    header = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 1:
            header = row
            break
    
    # 尺码列从第3列开始（索引2）
    size_cols = []
    for idx, val in enumerate(header):
        if idx >= 2 and val and str(val).strip() and str(val).strip() != 'None':
            size_cols.append((idx, str(val).strip()))
    
    # 读取数据
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        name = row[0]
        if not name or str(name).strip() == '':
            continue
        
        style_name = str(name).strip()
        size_weights = {}
        
        for col_idx, size_name in size_cols:
            weight = row[col_idx] if col_idx < len(row) else None
            if weight and str(weight).strip() and str(weight).strip() != '-':
                try:
                    size_weights[size_name] = float(weight)
                except:
                    pass
        
        if size_weights:
            weight_map[style_name] = size_weights
    
    return weight_map

weight_map_adult = build_weight_map('成人')
weight_map_kids = build_weight_map('儿童')

# 合并
all_weight_map = {}
all_weight_map.update(weight_map_kids)
all_weight_map.update(weight_map_adult)

print(f'重量表读取完成：儿童{len(weight_map_kids)}款 + 成人{len(weight_map_adult)}款 = 共{len(all_weight_map)}款')

# ===== 2. 读取款式库模板 =====
wb_style = load_workbook('款式库模板_v3_同步版2.xlsx')
ws_detail = wb_style['尺码明细']

# ===== 3. 填充净重 + 统计 =====
matched = 0
total = 0
unmatched_styles = set()

for row in ws_detail.iter_rows(min_row=2):
    total += 1
    style_name = str(row[1].value).strip() if row[1].value else ''
    size_name = str(row[3].value).strip() if row[3].value else ''
    
    if style_name in all_weight_map and size_name in all_weight_map[style_name]:
        weight = all_weight_map[style_name][size_name]
        row[4].value = weight  # 第5列是净重
        matched += 1
    else:
        unmatched_styles.add(style_name)

print(f'尺码明细共{total}行，成功匹配填充{matched}行')
if unmatched_styles:
    print(f'未匹配到重量的款式（{len(unmatched_styles)}个）:')
    for s in list(unmatched_styles)[:10]:
        print(f'  - {s}')

# ===== 4. 调整尺码明细字段顺序 =====
# 新表头：款式编号、款式名称、分类、尺码、净重(g)、衣长(cm)、肩宽(cm)、胸围(cm)、袖长(cm)、腰围(cm)、臀围(cm)、裤长(cm)、裤内长(cm)
new_headers = [
    '款式编号', '款式名称', '分类', '尺码',
    '净重(g)', '衣长(cm)', '肩宽(cm)', '胸围(cm)',
    '袖长(cm)', '腰围(cm)', '臀围(cm)', '裤长(cm)', '裤内长(cm)'
]

# 读取所有数据
all_rows = []
for row in ws_detail.iter_rows(min_row=2, values_only=True):
    all_rows.append(list(row))

# 重建sheet
wb_style.remove(ws_detail)
ws_new = wb_style.create_sheet('尺码明细', 1)  # 放到第2个位置

# 写入表头
ws_new.append(new_headers)
for col in range(1, len(new_headers) + 1):
    ws_new.cell(row=1, column=col).font = Font(bold=True)

# 写入数据（重新排列列）
# 原列索引: 0=编号, 1=名称, 2=分类, 3=尺码, 4=净重, 5=肩宽, 6=胸围, 7=衣长, 8=腰围, 9=裤长, 10=袖长, 11=其他
# 新列顺序: 编号, 名称, 分类, 尺码, 净重, 衣长, 肩宽, 胸围, 袖长, 腰围, 臀围(新), 裤长, 裤内长(新)
for row_data in all_rows:
    new_row = [
        row_data[0] if len(row_data) > 0 else '',  # 编号
        row_data[1] if len(row_data) > 1 else '',  # 名称
        row_data[2] if len(row_data) > 2 else '',  # 分类
        row_data[3] if len(row_data) > 3 else '',  # 尺码
        row_data[4] if len(row_data) > 4 else '',  # 净重
        row_data[7] if len(row_data) > 7 else '',  # 衣长（原第8列）
        row_data[5] if len(row_data) > 5 else '',  # 肩宽（原第6列）
        row_data[6] if len(row_data) > 6 else '',  # 胸围（原第7列）
        row_data[10] if len(row_data) > 10 else '', # 袖长（原第11列）
        row_data[8] if len(row_data) > 8 else '',  # 腰围（原第9列）
        '',  # 臀围（新增）
        row_data[9] if len(row_data) > 9 else '',  # 裤长（原第10列）
        '',  # 裤内长（新增）
    ]
    ws_new.append(new_row)

# 调整列宽
col_widths = [12, 36, 8, 10, 10, 12, 12, 12, 12, 12, 12, 12, 14]
for i, w in enumerate(col_widths, 1):
    ws_new.column_dimensions[chr(64 + i)].width = w

# 保存
output = '款式库模板_v4_带重量.xlsx'
wb_style.save(output)
print(f'\n已生成: {output}')
print('尺码明细字段顺序已调整为：衣长、肩宽、胸围、袖长、腰围、臀围、裤长、裤内长')
