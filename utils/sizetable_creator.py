"""
尺码表自动创建模块（供RPA上架流程调用）
"""
import time
from pathlib import Path
import openpyxl

EXCEL_PATH = Path(r"C:\Users\Administrator\.doubao\chats\2026-07-14\new-chat\temu_auto_publish_v2\款式库模板_v4_填尺码_最终版.xlsx")

# 参数名映射：Excel表头 → 妙手系统参数名
PARAM_MAP = {
    "衣长": "衣长",
    "胸围": "胸围",
    "肩宽": "肩宽",
    "袖长": "袖长",
    "裤长": "裤长",
    "腰围": "腰围",
    "臀围": "臀围",
    "裤内长": "裤内长",
    "建议身高": "建议身高",
    "建议体重": "建议体重",
}

def load_size_data(excel_path, style_name):
    """从Excel加载指定款式的尺码数据"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    # 找款式名所在行
    headers = None
    data_rows = []
    is_top = True  # 默认上衣
    for row in ws.iter_rows(values_only=True):
        if not row or not row[0]:
            continue
        if str(row[0]).strip() == "尺码":
            headers = [str(c).strip() if c else "" for c in row]
            continue
        if headers and str(row[0]).strip() == style_name:
            # 下一行开始是尺码数据
            break
    if not headers:
        return [], [], is_top
    # 判断是上衣还是裤子
    if "裤长" in headers or "腰围" in headers or "臀围" in headers:
        is_top = False
    # 读尺码数据
    for row in ws.iter_rows(values_only=True):
        if not row or not row[0]:
            continue
        size = str(row[0]).strip()
        if size in ["尺码", style_name] or not size.isdigit() and not size.startswith(("1", "2", "3", "4", "5", "6", "7", "8", "9", "S", "M", "L", "X")):
            continue
        # 只取有数据的行
        if any(c is not None and str(c).strip() for c in row[1:]):
            data_rows.append([size] + [str(c).strip() if c is not None else "" for c in row[1:len(headers)]])
    wb.close()
    return headers, data_rows, is_top

def map_param_labels(headers):
    """映射Excel表头到妙手参数名"""
    labels = []
    for h in headers[1:]:
        if not h:
            labels.append("")
            continue
        mapped = PARAM_MAP.get(h, h)
        labels.append(mapped)
    return labels

def get_size_list(data_rows):
    """获取尺码列表"""
    return [r[0] for r in data_rows]

def create_sizetable_for_style(page, style_name: str, cat_path: str = "") -> bool:
    """在已打开的页面中创建尺码表（page需已导航到尺码表管理页）"""
    print(f"  [DEBUG] 开始创建尺码表: {style_name}, 类目: {cat_path}")
    print(f"  [DEBUG] 当前页面URL: {page.url}")
    print(f"  [DEBUG] 当前页面标题: {page.title()}")
    headers, data_rows, is_top = load_size_data(EXCEL_PATH, style_name)
    print(f"  [DEBUG] 读取尺码数据: headers={len(headers) if headers else 0}列, data_rows={len(data_rows)}行")
    if not data_rows:
        print(f"  [SKIP] 无尺码数据")
        return False

    if not cat_path:
        print(f"  [SKIP] 无类目路径")
        return False

    param_labels = map_param_labels(headers)
    sizes = get_size_list(data_rows)
    data_map = {str(r[0]): r[1:] for r in data_rows}
    print(f"  创建尺码表: {style_name} 尺码:{sizes}")

    # 0. 打印页面所有按钮，调试用
    print(f"  [DEBUG] 页面所有按钮:")
    btns = page.locator("button").all()
    for i, btn in enumerate(btns):
        try:
            if btn.is_visible():
                txt = btn.inner_text().strip()
                if txt:
                    print(f"    [{i}] {txt}")
        except:
            pass

    # 0. 打开弹窗
    print(f"  [DEBUG] 查找创建按钮...")
    create_btn = page.get_by_role("button", name="创建尺码表模板")
    create_btn.wait_for(state="visible", timeout=15000)
    print(f"  [DEBUG] 找到创建按钮，点击...")
    create_btn.click()
    time.sleep(2)

    # 1. 名称
    page.get_by_role("textbox", name="*模板名称").click()
    page.get_by_role("textbox", name="*模板名称").fill(style_name)
    time.sleep(0.3)

    # 2. 类目 — 同step2_fill_only验证通过的逻辑
    cat_kw = cat_path.split("/")[-1].strip()
    page.get_by_role("textbox", name="*类目").click()
    time.sleep(0.3)
    page.get_by_role("textbox", name="*类目").fill(cat_kw)
    time.sleep(2)
    page.locator(f"li:has-text('{cat_kw}')").first.click(timeout=5000)
    time.sleep(1.5)

    # 3. 参数勾选（只勾有数据的列，默认勾选的空列要取消）
    active_params = []
    for j, pl in enumerate(param_labels):
        # 判断该列是否所有尺码都有数据
        has_data = False
        for r in data_rows:
            val = r[j+1]
            if val is not None and str(val).strip() != "":
                has_data = True
                break
        
        # JS精确控制勾选状态：有数据就勾，没数据就取消
        try:
            checked = page.evaluate(f"""(labelText) => {{
                const labels = document.querySelectorAll('label');
                for (const label of labels) {{
                    if ((label.innerText || '').trim() === labelText) {{
                        const cb = label.querySelector('.jx-checkbox__inner');
                        if (!cb) return null;
                        const isChecked = cb.parentElement.classList.contains('is-checked');
                        return isChecked;
                    }}
                }}
                return null;
            }}""", pl)
            
            if has_data:
                if checked is False:
                    # 没勾，点一下勾上
                    page.locator(f"label:has-text('{pl}') .jx-checkbox__inner").click(timeout=2000)
                    time.sleep(0.15)
                active_params.append(pl)
                print(f"    勾选参数: {pl}")
            else:
                if checked is True:
                    # 默认勾了但没数据，点一下取消
                    page.locator(f"label:has-text('{pl}') .jx-checkbox__inner").click(timeout=2000)
                    time.sleep(0.15)
                print(f"    跳过空列: {pl}")
        except Exception as e:
            if has_data:
                # 有数据的尽量尝试勾选
                try:
                    page.locator(f"label:has-text('{pl}') .jx-checkbox__inner").click(timeout=2000)
                    time.sleep(0.15)
                    active_params.append(pl)
                except:
                    print(f"    勾选参数失败: {pl}, {e}")
            else:
                print(f"    跳过空列: {pl}")
    param_labels = active_params  # 后续填充只用有数据的列
    time.sleep(1)

    # 4. 取消全选
    try:
        page.locator(".pro-virtual-table__checkbox.is-checked .jx-checkbox__inner").first.click(timeout=3000)
        time.sleep(0.3)
    except:
        pass

    # 5. 两阶段填充（同step2_fill_only验证过的逻辑）
    remaining = set(sizes)
    for i in range(80):
        page.evaluate("(p) => document.querySelector('.vue-recycle-scroller').scrollTop = p", i * 150)
        time.sleep(0.3)
        page.evaluate("""(args) => {
            const items = document.querySelectorAll('.vue-recycle-scroller__item-view');
            for (const item of items) {
                const txt = (item.innerText || '').trim().split(' ')[0];
                if (args.remaining.includes(txt)) {
                    const cb = item.querySelector('.jx-checkbox__inner');
                    if (cb) cb.click();
                }
            }
        }""", {"remaining": list(remaining)})
        time.sleep(0.15)

        result = page.evaluate("""(args) => {
            const filled = [];
            const items = document.querySelectorAll('.vue-recycle-scroller__item-view');
            for (const item of items) {
                const txt = (item.innerText || '').trim().split(' ')[0];
                if (args.remaining.includes(txt)) {
                    const inputs = item.querySelectorAll('input[type="text"]');
                    const cols = args.dataMap[txt];
                    if (cols && inputs.length > 0 && !inputs[0].disabled) {
                        for (let j = 0; j < cols.length && j < inputs.length; j++) {
                            const inp = inputs[j];
                            inp.focus();
                            inp.select();
                            document.execCommand('insertText', false, String(cols[j] || ''));
                        }
                        filled.push(txt);
                    }
                }
            }
            return filled;
        }""", {"remaining": list(remaining), "dataMap": data_map})

        for sz in result:
            remaining.discard(sz)
        if not remaining:
            break

    # 6. 保存
    page.get_by_role("button", name="保存").click(force=True)
    time.sleep(2)
    print(f"  [OK] 尺码表已创建: {style_name}")
    return True
