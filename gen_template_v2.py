"""
款式库模板 V2
- 款式名称以产品目录表为准
- 面料分类和克重分开两列
- 去掉单件净重参考
- 尺码明细表预留尺寸字段，数值全部留空
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import os
import re

os.chdir('C:/Users/Administrator/.doubao/chats/2026-07-14/new-chat/temu_auto_publish_v2')

# ===== 读取产品目录表（款式名称以此为准）=====
wb_cat = load_workbook('产品目录表.xlsx', data_only=True, read_only=True)

def parse_cat_sheet(sheet_name, category):
    ws = wb_cat[sheet_name]
    items = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        name = row[1]
        if not name or str(name).strip() in ['', '产品款名', '产品名称']:
            continue
        
        clean_name = str(name).strip().replace('\n', '')
        cost = row[5]
        fabric_raw = row[4]
        
        # 提取克重
        fabric_weight = ''
        if fabric_raw and str(fabric_raw).strip():
            g_match = re.search(r'(\d+\s*g)', str(fabric_raw), re.IGNORECASE)
            if g_match:
                fabric_weight = g_match.group(1).replace(' ', '')
        
        items.append({
            'name': clean_name,
            'category': category,
            'cost': cost,
            'fabric_weight': fabric_weight
        })
    return items

cat_kids = parse_cat_sheet('儿童款', '儿童')
cat_adult = parse_cat_sheet('成人', '成人')
all_items = cat_kids + cat_adult

# ===== 读取重量表获取尺码列表 =====
wb_weight = load_workbook('服装重量.xlsx', data_only=True)

def get_sizes(sheet_name):
    ws = wb_weight[sheet_name]
    sizes = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 1:
            for idx, val in enumerate(row):
                if idx >= 2 and val and str(val).strip() and str(val).strip() != 'None':
                    sizes.append(str(val).strip())
            break
    return sizes

sizes_kids = get_sizes('儿童')
sizes_adult = get_sizes('成人')

# ===== 生成Excel =====
wb_out = Workbook()

# Sheet1: 款式库主表
ws_main = wb_out.active
ws_main.title = '款式库主表'

headers = [
    '款式编号', '款式名称', '分类', '品类模板',
    '成本价(元)', '面料分类', '面料克重',
    '季节', '版型', '织造方式', '场景', '场合',
    '默认SKU分类', '单件毛重(g)',
    '尺码列表', '备注'
]
ws_main.append(headers)
for col in range(1, len(headers) + 1):
    ws_main.cell(row=1, column=col).font = Font(bold=True)

for idx, item in enumerate(all_items, 1):
    style_id = f"ST-{idx:03d}"
    sizes = sizes_kids if item['category'] == '儿童' else sizes_adult
    sizes_str = ','.join(sizes)
    
    ws_main.append([
        style_id, item['name'], item['category'], '',
        item['cost'], '', item['fabric_weight'],
        '', '', '', '', '',
        '', '',
        sizes_str, ''
    ])

# 列宽
widths = [12, 32, 8, 18, 12, 12, 12, 8, 8, 12, 8, 8, 12, 12, 36, 20]
for i, w in enumerate(widths, 1):
    ws_main.column_dimensions[chr(64 + i)].width = w

# Sheet2: 尺码明细
ws_detail = wb_out.create_sheet('尺码明细')
detail_headers = [
    '款式编号', '款式名称', '分类', '尺码',
    '净重(g)', '肩宽(cm)', '胸围(cm)', '衣长(cm)',
    '腰围(cm)', '裤长(cm)', '袖长(cm)', '其他'
]
ws_detail.append(detail_headers)
for col in range(1, len(detail_headers) + 1):
    ws_detail.cell(row=1, column=col).font = Font(bold=True)

for idx, item in enumerate(all_items, 1):
    style_id = f"ST-{idx:03d}"
    sizes = sizes_kids if item['category'] == '儿童' else sizes_adult
    for size in sizes:
        ws_detail.append([
            style_id, item['name'], item['category'], size,
            '', '', '', '', '', '', '', ''
        ])

detail_widths = [12, 32, 8, 10, 10, 12, 12, 12, 12, 12, 12, 15]
for i, w in enumerate(detail_widths, 1):
    ws_detail.column_dimensions[chr(64 + i)].width = w

# 保存
output = '款式库模板_v2.xlsx'
wb_out.save(output)
detail_count = sum(len(sizes_kids) if i['category'] == '儿童' else len(sizes_adult) for i in all_items)
print(f'生成完成: {output}')
print(f'主表: {len(all_items)} 款（儿童{len(cat_kids)} + 成人{len(cat_adult)}）')
print(f'尺码明细: {detail_count} 行（全部留空待填写）')
