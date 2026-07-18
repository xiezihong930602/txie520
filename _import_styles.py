# -*- coding: utf-8 -*-
"""导入款式名称到款式库表（去重）"""
import openpyxl
import subprocess
import json
import os

BASE_TOKEN = "Z69Pb8Zpua8h3PsKiumcidJ0nEg"
TABLE_ID = "tblgsBQC5kMCaDr4"
WORK_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_FILE = os.path.join(WORK_DIR, "款式库模板_v4_填尺码_最终版.xlsx")


def batch_create_records(records):
    payload = {"records": records}
    tmp_file = os.path.join(WORK_DIR, "_tmp_import.json")
    with open(tmp_file, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False)
    
    cmd = f'cmd /c "lark-cli base +record-batch-create --base-token {BASE_TOKEN} --table-id {TABLE_ID} --json @_tmp_import.json"'
    result = subprocess.run(cmd, cwd=WORK_DIR, capture_output=True, shell=True)
    output = result.stderr or result.stdout
    try:
        text = output.decode("utf-8", errors="ignore")
    except:
        text = output.decode("gbk", errors="ignore")
    try:
        resp = json.loads(text.strip())
    except:
        resp = {"ok": False, "error": {"message": text.strip()}}
    
    if os.path.exists(tmp_file):
        os.remove(tmp_file)
    return resp


def main():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    
    style_names = set()
    for row_idx in range(2, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=2).value
        if name:
            style_names.add(str(name).strip())
    
    records = [{"款式名称": name} for name in sorted(style_names)]
    print(f"共 {len(records)} 个款式，开始导入...")
    
    batch_size = 20
    success = 0
    for i in range(0, len(records), batch_size):
        batch = records[i:i+batch_size]
        resp = batch_create_records(batch)
        if resp.get("ok"):
            success += len(batch)
            print(f"  {i+1}-{min(i+batch_size, len(records))} 成功")
        else:
            print(f"  失败: {resp.get('error', {}).get('message', '未知错误')}")
    
    print(f"\n完成！成功导入 {success} 个款式名称")


if __name__ == "__main__":
    main()
