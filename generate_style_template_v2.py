"""
生成款式库Excel模板 V2
- 款式名称以产品目录表为准
- 面料分类和克重分开
- 尺码明细表预留尺寸字段（肩宽/衣长/裤长等）和重量，留空用户填
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import os
import re

os.chdir('C:/Users/Administrator/.doubao/chats/2026-07-14/new-chat/temu_auto_publish_v2')

# 读取产品目录表（款式名称以此为准）
wb_cat = load_workbook('产品目录表.xlsx', data_only=True, read_only=True)

def parse_cat_sheet(sheet_name, category):
    """解析产品目录表"""
    ws = wb_cat[sheet_name]
    items = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        name = row[1]
        if not name or str(name).strip() in ['', '产品款名', '产品名称']:
            continue
        
        clean_name = str(name).strip().replace('\n', '')
        cost = row[5]
        fabric_raw = row[4]  # 原始面料信息，如"180g拉毛"
        
        # 解析面料分类和克重
        fabric_type = ''
        fabric_weight = ''
        if fabric_raw and str(fabric_raw).strip():
            fabric_str = str(fabric_raw).strip()
            # 提取克重（数字+g）
            g_match = re.search(r'(\d+)\s*g', fabric_str, re.IGNORECASE)
            if g_match:
                fabric_weight = g_match.group(1) + 'g'
            # 剩下的是面料类型描述
            # 简单处理，实际面料分类（针织/梭织）需要人工填
            fabric_type = ''  # 留空，用户填针织/梭织
        
        items.append({
            'name': clean_name,
            'category': category,
            'cost': cost,
            'fabric_raw': fabric_raw,
            'fabric_type': fabric_type,
            'fabric_weight': fabric_weight
        })
    return items

cat_kids = parse_cat_sheet('儿童款', '儿童')
cat_adult = parse_cat_sheet('成人', '成人')
all_items = cat_kids + cat_adult

print(f'产品目录表共 {len(all_items)} 款（儿童{len(cat_kids)} + 成人{len(cat_adult)}）')

# 读取重量表的尺码信息（只提取尺码列表，重量留空）
wb_weight = load_workbook('服装重量.xlsx', data_only=True)

def get_size_list(sheet_name):
    """获取重量表中的尺码列表"""
    ws = wb_weight[sheet_name]
    sizes = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 1:  # 第2行是表头
            for idx, val in enumerate(row):
                if idx >= 2 and val and str(val).strip() and str(val).strip() != 'None':
                    sizes.append(str(val).strip())
            break
    return sizes

sizes_kids = get_size_list('儿童')  # 儿童款通用尺码
sizes_adult = get_size_list('成人')  # 成人款通用尺码

# 生成Excel
wb_out = Workbook()

# ========== Sheet1: 款式库主表 ==========
ws_main = wb_out.active
ws_main.title = '款式库主表'

headers = [
    '款式编号', '款式名称', '分类', '品类模板',
    '成本价(元)', '面料分类', '面料克重',
    '季节', '版型', '织造方式', '场景', '场合',
    '默认SKU分类', '单件毛重(g)',
    '尺码列表', '备注'
]
ws_main.append(headers)
for col in range(1, len(headers) + 1):
    ws_main.cell(row=1, column=col).font = Font(bold=True)

# 填充款式数据
for idx, item in enumerate(all_items, 1):
    style_id = f"ST-{idx:03d}"
    
    # 尺码列表（用对应分类的通用尺码）
    sizes = sizes_kids if item['category'] == '儿童' else sizes_adult
    sizes_str = ','.join(sizes)
    
    row_data = [
        style_id,
        item['name'],
        item['category'],
        '',  # 品类模板
        item['cost'],
        item['fabric_type'],  # 面料分类（针织/梭织），留空用户填
        item['fabric_weight'],  # 克重，能提取的自动填
        '',  # 季节
        '',  # 版型
        '',  # 织造方式
        '',  # 场景
        '',  # 场合
        '',  # SKU分类
        '',  # 毛重
        sizes_str,
        ''   # 备注
    ]
    ws_main.append(row_data)

# 调整列宽
col_widths = [12, 32, 8, 18, 12, 12, 12, 8, 8, 12, 8, 8, 12, 12, 36, 20]
for i, w in enumerate(col_widths, 1):
    ws_main.column_dimensions[chr(64 + i)].width = w

# ========== Sheet2: 尺码明细表 ==========
ws_detail = wb_out.create_sheet('尺码明细')

detail_headers = [
    '款式编号', '款式名称', '分类', '尺码',
    '净重(g)', '肩宽(cm)', '胸围(cm)', '衣长(cm)',
    '腰围(cm)', '裤长(cm)', '其他尺寸'
]
ws_detail.append(detail_headers)
for col in range(1, len(detail_headers) + 1):
    ws_detail.cell(row=1, column=col).font = Font(bold=True)

# 每个款式每个尺码生成一行（数值全部留空，用户填）
for idx, item in enumerate(all_items, 1):
    style_id = f"ST-{idx:03d}"
    sizes = sizes_kids if item['category'] == '儿童' else sizes_adult
    
    for size in sizes:
        ws_detail.append([
            style_id,
            item['name'],
            item['category'],
            size,
            '',  # 净重
            '',  # 肩宽
            '',  # 胸围
            '',  # 衣长
            '',  # 腰围
            '',  # 裤长
            ''   # 其他
        ])

# 调整列宽
detail_widths = [12, 32, 8, 10, 10, 12, 12, 12, 12, 12, 15]
for i, w in enumerate(detail_widths, 1):
    ws_detail.column_dimensions[chr(64 + i)].width = w

# 保存
output_file = '款式库模板_v2.xlsx'
wb_out.save(output_file)
detail_rows = sum(len(sizes_kids) if c == '儿童' else len(sizes_adult) for c in [i['category'] for i in all_items])
print(f'已生成 {output_file}')
print(f'主表 {len(all_items)} 行，尺码明细 {detail_rows} 行')
print('说明：尺码明细中的尺寸数据和重量全部留空，请自行填写')
