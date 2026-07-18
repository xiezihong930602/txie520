# 从Excel加载款式库数据
import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook
from models.style import Style
from typing import List, Dict


def load_styles_from_excel(file_path: str) -> List[Style]:
    """
    从款式库Excel加载所有款式数据
    
    主表字段：款式编号、款式名称、分类、成本价(元)、面料克重、季节、版型、织造方式、尺码列表、备注
    尺码明细表：款式编号、款式名称、分类、尺码、净重(g)、衣长(cm)、肩宽(cm)、胸围(cm)、袖长(cm)、腰围(cm)、臀围(cm)、裤长(cm)、裤内长(cm)
    """
    wb = load_workbook(file_path, data_only=True)
    
    # 主表（第一个sheet）
    main_sheet = wb.worksheets[0]
    headers = [cell.value for cell in main_sheet[1]]
    col_map = {h: i for i, h in enumerate(headers) if h}
    
    styles = []
    for row_idx in range(2, main_sheet.max_row + 1):
        row = main_sheet[row_idx]
        
        def _get(col_name, default=""):
            idx = col_map.get(col_name)
            if idx is None or idx >= len(row):
                return default
            val = row[idx].value
            return val if val is not None else default
        
        style_id = _get("款式编号")
        style_name = _get("款式名称")
        
        if not style_id or not style_name:
            continue
        
        cost_val = _get("成本价(元)", 0)
        try:
            cost_price = float(cost_val) if cost_val else 0.0
        except:
            cost_price = 0.0
        
        style = Style(
            style_id=str(style_id),
            style_name=str(style_name),
            template_name="",  # 款式库不存模板名，上架时指定
            category=str(_get("分类", "")),
            fabric_weight=str(_get("面料克重", "")),
            season=str(_get("季节", "")),
            fit=str(_get("版型", "")),
            weave_method=str(_get("织造方式", "")),
            cost_price=cost_price,
            sizes=[],  # 后面从尺码明细表填充
            remark=str(_get("备注", "")) if _get("备注") else None,
        )
        styles.append(style)
    
    # 读取尺码明细表（第二个sheet）
    if len(wb.worksheets) >= 2:
        size_sheet = wb.worksheets[1]
        size_headers = [cell.value for cell in size_sheet[1]]
        size_col_map = {h: i for i, h in enumerate(size_headers) if h}
        
        style_sizes: Dict[str, list] = {}
        style_details: Dict[str, dict] = {}
        style_weights: Dict[str, float] = {}  # 取M码/中间码作为默认净重
        
        for row_idx in range(2, size_sheet.max_row + 1):
            row = size_sheet[row_idx]
            
            def _sget(col_name, default=None):
                idx = size_col_map.get(col_name)
                if idx is None or idx >= len(row):
                    return default
                return row[idx].value
            
            s_name = _sget("款式名称")
            size_val = _sget("尺码")
            
            if not s_name or not size_val:
                continue
            
            s_name = str(s_name)
            size_val = str(size_val)
            
            if s_name not in style_sizes:
                style_sizes[s_name] = []
                style_details[s_name] = {}
            
            style_sizes[s_name].append(size_val)
            
            # 读取尺码明细字段
            detail = {}
            for field in ["净重(g)", "衣长(cm)", "肩宽(cm)", "胸围(cm)", "袖长(cm)",
                         "腰围(cm)", "臀围(cm)", "裤长(cm)", "裤内长(cm)"]:
                val = _sget(field)
                if val is not None:
                    try:
                        detail[field] = float(val)
                    except:
                        pass
            style_details[s_name][size_val] = detail
            
            # 记录净重
            net_w = _sget("净重(g)")
            if net_w:
                try:
                    style_weights[s_name] = float(net_w)  # 先存第一个，后面再取中间值
                except:
                    pass
        
        # 回填到style对象
        name_map = {s.style_name: s for s in styles}
        for s_name, sizes in style_sizes.items():
            if s_name in name_map:
                sorted_sizes = sorted(sizes, key=_size_sort_key)
                name_map[s_name].sizes = sorted_sizes
                name_map[s_name].size_details = style_details.get(s_name, {})
                # 默认净重取中间尺码的
                if sorted_sizes and s_name in style_details:
                    mid_idx = len(sorted_sizes) // 2
                    mid_size = sorted_sizes[mid_idx]
                    mid_detail = style_details[s_name].get(mid_size, {})
                    if "净重(g)" in mid_detail:
                        name_map[s_name].net_weight = mid_detail["净重(g)"]
                        name_map[s_name].gross_weight = mid_detail["净重(g)"] * 1.1  # 毛重估算
    
    wb.close()
    return styles


def _size_sort_key(size: str):
    """尺码排序：数字尺码按数值排，字母尺码按固定顺序排"""
    size_upper = str(size).upper().strip()
    # 纯数字尺码（如120, 130）
    if size_upper.isdigit():
        return (0, int(size_upper))
    # 字母尺码
    letter_order = {
        "XS": 0, "S": 1, "M": 2, "L": 3, 
        "XL": 4, "2XL": 5, "3XL": 6, "4XL": 7,
        "XXL": 5, "XXXL": 6
    }
    return (1, letter_order.get(size_upper, 99), size_upper)


def get_style_by_name(styles: List[Style], name: str) -> Style | None:
    """按款式名称查找"""
    for s in styles:
        if s.style_name == name:
            return s
    return None


def get_style_by_id(styles: List[Style], style_id: str) -> Style | None:
    """按款式编号查找"""
    for s in styles:
        if s.style_id == style_id:
            return s
    return None


if __name__ == "__main__":
    # 测试
    file = r"C:\Users\Administrator\.doubao\chats\2026-07-14\new-chat\temu_auto_publish_v2\款式库模板_v4_填尺码_最终版.xlsx"
    styles = load_styles_from_excel(file)
    print(f"加载款式数量: {len(styles)}")
    print()
    for s in styles[:3]:
        print(f"【{s.style_id}】{s.style_name}")
        print(f"  分类: {s.category}")
        print(f"  成本价: {s.cost_price}")
        print(f"  面料克重: {s.fabric_weight}")
        print(f"  季节: {s.season}")
        print(f"  版型: {s.fit}")
        print(f"  织造方式: {s.weave_method}")
        print(f"  尺码: {s.sizes}")
        print(f"  净重(中间码): {s.net_weight}g")
        if s.size_details:
            first_size = list(s.size_details.keys())[0]
            print(f"  样例尺码{first_size}明细: {s.size_details[first_size]}")
        print()
