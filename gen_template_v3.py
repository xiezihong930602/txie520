"""
款式库模板 V3 - 优化版
- 套装拆分为独立款式（上衣+裤子）
- 尺码范围展开为具体尺码
- 精简字段（去掉品类模板、场景、场合、SKU分类、毛重、面料分类）
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
import os
import re

os.chdir('C:/Users/Administrator/.doubao/chats/2026-07-14/new-chat/temu_auto_publish_v2')

# 读取用户填写版
wb_in = load_workbook('款式库模板_用户填写版.xlsx', data_only=True)
ws_in = wb_in['款式库主表']

def parse_cost(cost_str):
    """解析成本价，返回(上衣价格, 裤子价格)。如果是单件，裤子价格为None"""
    if not cost_str or str(cost_str).strip() == '':
        return None, None
    
    s = str(cost_str).strip()
    
    # 尝试匹配衣裤分开的格式
    # 格式1: 衣：16 裤：15
    # 格式2: 衣服：10.5  裤12
    # 格式3: 衣：8.5   裤：7.5
    top_match = re.search(r'衣[服]?[：:]\s*([\d.]+)', s)
    bot_match = re.search(r'裤[子]?[：:]?\s*([\d.]+)', s)
    
    if top_match and bot_match:
        return float(top_match.group(1)), float(bot_match.group(1))
    
    # 单个数字，就是单件
    try:
        val = float(s)
        return val, None
    except:
        pass
    
    return s, None  # 解析不了原样返回

def parse_sizes(size_str):
    """解析尺码字符串，返回尺码列表"""
    if not size_str or str(size_str).strip() == '' or str(size_str).strip() == 'None':
        return []
    
    s = str(size_str).strip()
    
    # 逗号分隔的直接返回
    if ',' in s:
        return [x.strip() for x in s.split(',') if x.strip()]
    
    # 范围格式：120-160, 130-170, 90-130
    range_match = re.match(r'^(\d+)\s*-\s*(\d+)$', s)
    if range_match:
        start = int(range_match.group(1))
        end = int(range_match.group(2))
        step = 10
        sizes = []
        for v in range(start, end + 1, step):
            sizes.append(str(v))
        return sizes
    
    # ST-033特殊格式：160165170175（连续数字，每3位一个尺码）
    if re.match(r'^\d{12}$', s):  # 4个3位数字
        return [s[i:i+3] for i in range(0, 12, 3)]
    if re.match(r'^\d{9,}$', s):  # 其他连续数字
        # 尝试每3位拆分
        sizes = []
        for i in range(0, len(s), 3):
            if i + 3 <= len(s):
                sizes.append(s[i:i+3])
        if sizes:
            return sizes
    
    # 单个尺码
    return [s]

# 读取所有款式
raw_styles = []
for i, row in enumerate(ws_in.iter_rows(values_only=True)):
    if i == 0:
        continue
    if not row[0]:
        continue
    
    raw_styles.append({
        'id': row[0],
        'name': str(row[1]).strip() if row[1] else '',
        'category': str(row[2]).strip() if row[2] else '',
        'cost_raw': row[4],
        'fabric_weight': str(row[6]).strip() if row[6] and str(row[6]).strip() else '',
        'season': str(row[7]).strip() if row[7] and str(row[7]).strip() else '',
        'version': str(row[8]).strip() if row[8] and str(row[8]).strip() else '',
        'weave': str(row[9]).strip() if row[9] and str(row[9]).strip() else '',
        'sizes_raw': row[14],
        'remark': str(row[15]).strip() if row[15] and str(row[15]).strip() else ''
    })

# 拆分套装
all_styles = []
new_id = 0

for style in raw_styles:
    top_cost, bot_cost = parse_cost(style['cost_raw'])
    sizes = parse_sizes(style['sizes_raw'])
    
    # 判断是否套装
    is_suit = bot_cost is not None and bot_cost != ''
    
    if is_suit:
        # 拆成上衣和裤子，智能命名
        original_name = style['name']
        
        # ===== 上衣名称 =====
        top_name = original_name
        # 去掉裤子相关的后缀
        top_name = re.sub(r'[+及和与].*[裤装]$', '', top_name)
        top_name = re.sub(r'卫裤.*$', '', top_name)
        top_name = re.sub(r'短裤.*$', '', top_name)
        top_name = re.sub(r'长裤.*$', '', top_name)
        top_name = re.sub(r'裤子.*$', '', top_name)
        top_name = re.sub(r'套装$', '', top_name)
        top_name = top_name.strip(' 及+与')
        
        # 如果名称里没有明显的上衣标识，加上衣
        if not any(k in top_name for k in ['卫衣', 'T恤', '短袖', '背心', '衫', '夹克', '上衣']):
            top_name = top_name + '上衣'
        
        # ===== 裤子名称 =====
        bot_name = original_name
        # 提取裤子类型
        pant_type = '裤子'
        if '卫裤' in original_name:
            pant_type = '卫裤'
        elif '短裤' in original_name:
            pant_type = '短裤'
        elif '长裤' in original_name:
            pant_type = '长裤'
        
        # 去掉上衣相关前缀
        bot_name = re.sub(r'^.*?[+及和与]', '', bot_name)
        if bot_name == original_name:  # 没匹配到分隔符，尝试去掉开头的上衣词
            bot_name = re.sub(r'^.*?卫衣', '', bot_name)
            bot_name = re.sub(r'^.*?T恤', '', bot_name)
            bot_name = re.sub(r'^.*?短袖', '', bot_name)
            bot_name = re.sub(r'^.*?背心', '', bot_name)
        
        if not bot_name or bot_name.strip() in ['', pant_type]:
            # 提取不到，用面料/风格 + 裤子类型
            prefix = re.sub(r'(卫衣|T恤|短袖|背心|衫|夹克|上衣).*$', '', original_name).strip()
            bot_name = prefix + pant_type
        else:
            bot_name = bot_name.strip()
            if not any(k in bot_name for k in ['裤']):
                bot_name = bot_name + pant_type
        
        # 兜底：如果名字还是不对，用原名 + 裤子
        if len(bot_name) < 2:
            bot_name = original_name + '（裤子）'
        
        # 上衣
        new_id += 1
        all_styles.append({
            'id': f'ST-{new_id:03d}',
            'name': top_name,
            'category': style['category'],
            'cost': top_cost,
            'fabric_weight': style['fabric_weight'],
            'season': style['season'],
            'version': style['version'],
            'weave': style['weave'],
            'sizes': sizes,
            'remark': style['remark'],
            'type': '上衣'
        })
        
        # 裤子
        new_id += 1
        all_styles.append({
            'id': f'ST-{new_id:03d}',
            'name': bot_name,
            'category': style['category'],
            'cost': bot_cost,
            'fabric_weight': style['fabric_weight'],
            'season': style['season'],
            'version': style['version'],
            'weave': style['weave'],
            'sizes': sizes,
            'remark': style['remark'],
            'type': '裤子'
        })
    else:
        # 单件
        new_id += 1
        all_styles.append({
            'id': f'ST-{new_id:03d}',
            'name': style['name'],
            'category': style['category'],
            'cost': top_cost if top_cost else style['cost_raw'],
            'fabric_weight': style['fabric_weight'],
            'season': style['season'],
            'version': style['version'],
            'weave': style['weave'],
            'sizes': sizes,
            'remark': style['remark'],
            'type': '单件'
        })

print(f'拆分后共 {len(all_styles)} 款')

# ===== 生成Excel =====
wb_out = Workbook()

# Sheet1: 款式库主表
ws_main = wb_out.active
ws_main.title = '款式库主表'

headers = [
    '款式编号', '款式名称', '分类',
    '成本价(元)', '面料克重',
    '季节', '版型', '织造方式',
    '尺码列表', '备注'
]
ws_main.append(headers)
for col in range(1, len(headers) + 1):
    ws_main.cell(row=1, column=col).font = Font(bold=True)

for s in all_styles:
    sizes_str = ','.join(s['sizes']) if s['sizes'] else ''
    ws_main.append([
        s['id'], s['name'], s['category'],
        s['cost'], s['fabric_weight'],
        s['season'], s['version'], s['weave'],
        sizes_str, s['remark']
    ])

# 列宽
widths = [12, 36, 8, 12, 12, 8, 8, 12, 36, 20]
for i, w in enumerate(widths, 1):
    ws_main.column_dimensions[chr(64 + i)].width = w

# Sheet2: 尺码明细
ws_detail = wb_out.create_sheet('尺码明细')
detail_headers = [
    '款式编号', '款式名称', '分类', '尺码',
    '净重(g)', '肩宽(cm)', '胸围(cm)', '衣长(cm)',
    '腰围(cm)', '裤长(cm)', '袖长(cm)', '其他'
]
ws_detail.append(detail_headers)
for col in range(1, len(detail_headers) + 1):
    ws_detail.cell(row=1, column=col).font = Font(bold=True)

detail_count = 0
for s in all_styles:
    for size in s['sizes']:
        ws_detail.append([
            s['id'], s['name'], s['category'], size,
            '', '', '', '', '', '', '', ''
        ])
        detail_count += 1

detail_widths = [12, 36, 8, 10, 10, 12, 12, 12, 12, 12, 12, 15]
for i, w in enumerate(detail_widths, 1):
    ws_detail.column_dimensions[chr(64 + i)].width = w

# 保存
output = '款式库模板_v3.xlsx'
wb_out.save(output)
print(f'生成完成: {output}')
print(f'主表: {len(all_styles)} 款')
print(f'尺码明细: {detail_count} 行（全部留空）')
print()
print('前10款预览:')
for s in all_styles[:10]:
    print(f"  {s['id']} | {s['name']} | {s['cost']}元 | {','.join(s['sizes'][:3])}...")
