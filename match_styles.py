from openpyxl import load_workbook
import os

os.chdir('C:/Users/Administrator/.doubao/chats/2026-07-14/new-chat/temu_auto_publish_v2')

# 读取重量表款式
wb1 = load_workbook('服装重量.xlsx', data_only=True)

def get_styles(sheet_name):
    ws = wb1[sheet_name]
    styles = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        name = row[0]
        if name and str(name).strip():
            styles.append(str(name).strip())
    return styles

weight_kids = get_styles('儿童')
weight_adult = get_styles('成人')

# 读取产品目录表款式
wb2 = load_workbook('产品目录表.xlsx', data_only=True, read_only=True)

def get_cat_styles(sheet_name):
    ws = wb2[sheet_name]
    styles = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue
        name = row[1]
        if name and str(name).strip() and str(name).strip() != '产品款名':
            styles.append({
                'name': str(name).strip().replace('\n', ''),
                'cost': row[5],
                'fabric': row[4]
            })
    return styles

cat_kids = get_cat_styles('儿童款')
cat_adult = get_cat_styles('成人')

# 简单匹配：计算字符相似度
def simple_match(weight_names, cat_list):
    matches = []
    used_cat = set()
    
    for w_name in weight_names:
        matched = None
        best_score = 0
        best_idx = -1
        
        w_clean = w_name.replace(' ', '')
        
        for j, cat in enumerate(cat_list):
            if j in used_cat:
                continue
            c_clean = cat['name'].replace(' ', '')
            # 计算共同字符比例
            common = len(set(w_clean) & set(c_clean))
            score = common / max(len(w_clean), len(c_clean))
            
            if score > best_score:
                best_score = score
                best_idx = j
        
        if best_score > 0.4:  # 相似度阈值
            matched = cat_list[best_idx]
            used_cat.add(best_idx)
        
        matches.append({
            'weight_name': w_name,
            'cat_name': matched['name'] if matched else '【未匹配】',
            'cost': matched['cost'] if matched else '',
            'fabric': matched['fabric'] if matched else '',
            'score': round(best_score, 2)
        })
    
    unmatched_cat = [cat for j, cat in enumerate(cat_list) if j not in used_cat]
    return matches, unmatched_cat

# 儿童款匹配
matches_kids, unmatched_kids = simple_match(weight_kids, cat_kids)

# 成人款匹配
matches_adult, unmatched_adult = simple_match(weight_adult, cat_adult)

# 输出结果
result = []
result.append('=== 儿童款匹配对照 ===')
matched_count = len([m for m in matches_kids if m['cat_name'] != '【未匹配】'])
result.append(f'重量表 {len(weight_kids)} 款 | 目录表 {len(cat_kids)} 款 | 匹配上 {matched_count} 款')
result.append('')
result.append('重量表款式 → 目录表款式 | 成本 | 面料 | 相似度')
result.append('-' * 70)
for m in matches_kids:
    line = f"{m['weight_name']}  →  {m['cat_name']}  | {m['cost']} | {m['fabric']} | {m['score']}"
    result.append(line)

if unmatched_kids:
    result.append(f'\n【目录表未匹配上的款式】共 {len(unmatched_kids)} 款')
    for cat in unmatched_kids:
        result.append(f"  - {cat['name']} | 成本: {cat['cost']} | 面料: {cat['fabric']}")

result.append('\n\n=== 成人款匹配对照 ===')
matched_count = len([m for m in matches_adult if m['cat_name'] != '【未匹配】'])
result.append(f'重量表 {len(weight_adult)} 款 | 目录表 {len(cat_adult)} 款 | 匹配上 {matched_count} 款')
result.append('')
result.append('重量表款式 → 目录表款式 | 成本 | 面料 | 相似度')
result.append('-' * 70)
for m in matches_adult:
    line = f"{m['weight_name']}  →  {m['cat_name']}  | {m['cost']} | {m['fabric']} | {m['score']}"
    result.append(line)

if unmatched_adult:
    result.append(f'\n【目录表未匹配上的款式】共 {len(unmatched_adult)} 款')
    for cat in unmatched_adult:
        result.append(f"  - {cat['name']} | 成本: {cat['cost']} | 面料: {cat['fabric']}")

with open('款式匹配对照表.txt', 'w', encoding='utf-8') as f:
    f.write('\n'.join(result))

print('已生成 款式匹配对照表.txt')
